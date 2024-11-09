from data_prepare import true_results, model_predictions
import numpy as np
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix
from openpyxl import Workbook
import os
# 获取当前文件夹下的路径
pwd = os.path.dirname(os.path.abspath(__file__))

print("真实结果：", true_results)
print("模型预测结果：", model_predictions)


def evaluate_model_performance(true_results, model_predictions):
    model_metrics = {}

    # 遍历每个模型
    for model_name, predictions in model_predictions.items():
        accuracies = []
        precisions = []
        recalls = []
        f1_scores = []
        conf_matrices = []
        sample_metrics = []  # 用于保存每个样本的指标

        # 遍历每个样本
        for true, pred in zip(true_results, predictions):
            # 将列表转换为 numpy 数组
            y_true = np.array(true)
            y_pred = np.array(pred)

            # 计算各项指标
            acc = accuracy_score(y_true, y_pred)
            precision = precision_score(y_true, y_pred, average='binary', pos_label=1)
            recall = recall_score(y_true, y_pred, average='binary', pos_label=1)
            f1 = f1_score(y_true, y_pred, average='binary', pos_label=1)
            conf_matrix = confusion_matrix(y_true, y_pred, labels=[1, 2])

            # 保存单个样本的结果
            sample_metrics.append({
                'accuracy': acc,
                'precision': precision,
                'recall': recall,
                'f1_score': f1,
                'confusion_matrix': conf_matrix
            })

            # 保存整体统计
            accuracies.append(acc)
            precisions.append(precision)
            recalls.append(recall)
            f1_scores.append(f1)
            conf_matrices.append(conf_matrix)

        # 计算平均指标
        avg_accuracy = np.mean(accuracies)
        avg_precision = np.mean(precisions)
        avg_recall = np.mean(recalls)
        avg_f1 = np.mean(f1_scores)

        # 保存模型的总体指标和每个样本的指标
        model_metrics[model_name] = {
            'overall_metrics': {
                'accuracy': avg_accuracy,
                'precision': avg_precision,
                'recall': avg_recall,
                'f1_score': avg_f1
            },
            'sample_metrics': sample_metrics  # 每个样本的具体指标
        }

    return model_metrics


# 调用函数并输出结果
model_metrics = evaluate_model_performance(true_results, model_predictions)

# 输出每个模型的结果
for model_name, metrics in model_metrics.items():
    print(f"模型 {model_name} 的平均指标：")
    print(f"准确率（Accuracy）：{metrics['overall_metrics']['accuracy']:.2%}")
    print(f"精确率（Precision）：{metrics['overall_metrics']['precision']:.2%}")
    print(f"召回率（Recall）：{metrics['overall_metrics']['recall']:.2%}")
    print(f"F1-score：{metrics['overall_metrics']['f1_score']:.2%}")
    print("-" * 30)

    # 输出每个样本的指标
    print(f"模型 {model_name} 每个样本的指标：")
    for idx, sample_metric in enumerate(metrics['sample_metrics'], start=1):
        print(f"样本 {idx}：")
        print(f"  准确率（Accuracy）：{sample_metric['accuracy']:.2%}")
        print(f"  精确率（Precision）：{sample_metric['precision']:.2%}")
        print(f"  召回率（Recall）：{sample_metric['recall']:.2%}")
        print(f"  F1-score：{sample_metric['f1_score']:.2%}")
        print(f"  混淆矩阵（Confusion Matrix）：\n{sample_metric['confusion_matrix']}")
        print("-" * 20)

# 初始化工作簿
wb = Workbook()


# 定义函数，将模型的预测结果、指标以及平均预测指标输出到 Excel
def output_predictions_to_excel(true_results, model_predictions, workbook):
    num_questions = len(true_results[0])  # 每个样本的问题数量
    model_average_metrics = {model_name: {'accuracy': [], 'precision': [], 'recall': [], 'f1': []}
                             for model_name in model_predictions.keys()}
    question_avg_stats = [{'TP': 0, 'FP': 0, 'FN': 0, 'TN': 0} for _ in range(num_questions)]
    num_samples = len(true_results)

    # 遍历每个样本，创建独立工作表
    for sample_idx, true in enumerate(true_results, start=1):
        # 创建一个新的工作表
        sheet = workbook.create_sheet(title=f"Sample_{sample_idx}")

        # 添加列标签（Q1, Q2, ...）
        sheet.append([""] + [f"Q{i + 1}" for i in range(num_questions)])

        # 添加真实值行
        sheet.append(["True"] + true)

        # 添加每个模型的预测结果行，并计算每个题项的混淆矩阵
        for model_name, predictions in model_predictions.items():
            pred = predictions[sample_idx - 1]
            sheet.append([model_name] + pred)  # 针对当前样本的预测结果

        # 添加性能指标
        sheet.append([""])
        sheet.append(["模型性能指标："])
        for model_name, predictions in model_predictions.items():
            pred = predictions[sample_idx - 1]
            y_true = np.array(true)
            y_pred = np.array(pred)
            accuracy = accuracy_score(y_true, y_pred)
            precision = precision_score(y_true, y_pred, average='binary', pos_label=1)
            recall = recall_score(y_true, y_pred, average='binary', pos_label=1)
            f1 = f1_score(y_true, y_pred, average='binary', pos_label=1)
            sheet.append(
                [model_name, f"Accuracy: {accuracy:.2%}", f"Precision: {precision:.2%}", f"Recall: {recall:.2%}",
                 f"F1: {f1:.2%}"])

            # 更新模型的平均指标
            model_average_metrics[model_name]['accuracy'].append(accuracy)
            model_average_metrics[model_name]['precision'].append(precision)
            model_average_metrics[model_name]['recall'].append(recall)
            model_average_metrics[model_name]['f1'].append(f1)

        # 添加空行以美化布局
        sheet.append([""])


        # 平均准确率、精确率、召回率和 F1-score
        accuracy_avg = []
        precision_avg = []
        recall_avg = []
        f1_avg = []


        # 写入平均指标
        sheet.append(["Accuracy"] + [f"{x:.2%}" for x in accuracy_avg])
        sheet.append(["Precision"] + [f"{x:.2%}" for x in precision_avg])
        sheet.append(["Recall"] + [f"{x:.2%}" for x in recall_avg])
        sheet.append(["F1-score"] + [f"{x:.2%}" for x in f1_avg])

    # 删除默认的 Sheet 工作表（如果存在）
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    # 生成 `Average` 工作表
    average_sheet = workbook.create_sheet(title="Average")
    average_sheet.append(["模型总体平均指标"])
    average_sheet.append(["模型名称", "平均准确率", "平均精确率", "平均召回率", "平均F1分数"])
    for model_name, metrics in model_average_metrics.items():
        avg_accuracy = np.mean(metrics['accuracy'])
        avg_precision = np.mean(metrics['precision'])
        avg_recall = np.mean(metrics['recall'])
        avg_f1 = np.mean(metrics['f1'])
        average_sheet.append(
            [model_name, f"{avg_accuracy:.2%}", f"{avg_precision:.2%}", f"{avg_recall:.2%}", f"{avg_f1:.2%}"])


# 调用函数将数据写入 Excel 文件
output_predictions_to_excel(true_results, model_predictions, wb)

# 保存 Excel 文件
wb.save(f"{pwd}/model_predictions_comparison.xlsx")