import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from data_prepare import true_results, model_predictions, sample_ids
plt.rcParams['font.sans-serif'] = ['SimHei']  # 用黑体显示中文
plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import os
pwd = os.path.dirname(os.path.realpath(__file__))
# 假设模型名称列表
model_names = list(model_predictions.keys())
num_samples = len(sample_ids)
num_models = len(model_names)

# 假设模型名称列表
model_names = list(model_predictions.keys())
num_samples = len(sample_ids)
num_models = len(model_names)

# 创建数据点列表和对应的标签
data_points = []
labels = []

for i in range(num_samples):
    # 添加真实结果
    true_result = true_results[i]
    data_points.append(true_result)
    labels.append(f'sample{i+1}_true')
    # 添加每个模型的预测结果
    for model_name in model_names:
        prediction = model_predictions[model_name][i]
        data_points.append(prediction)
        labels.append(f'sample{i+1}_{model_name}')

# 总数据点数量
num_data_points = len(data_points)

# 初始化相似度矩阵
similarity_matrix = np.zeros((num_data_points, num_data_points))

# 计算两两相似度
for i in range(num_data_points):
    for j in range(num_data_points):
        data_i = data_points[i]
        data_j = data_points[j]
        # 计算预测结果中相同的数量
        agree = sum(p == q for p, q in zip(data_i, data_j))
        total = len(data_i)
        # 计算相似性（相同数量占总数量的比例）
        similarity = agree / total
        similarity_matrix[i, j] = similarity

# 使用Seaborn绘制热力图（可选）
plt.figure(figsize=(12, 10))
sns.heatmap(similarity_matrix, annot=False, cmap='YlGnBu',
            xticklabels=labels, yticklabels=labels)
plt.title('数据点之间的相似度热力图')
plt.xlabel('数据点')
plt.ylabel('数据点')
plt.xticks(rotation=90)
plt.tight_layout()
# 保存热力图到文件
plt.savefig('data_point_similarity_heatmap.png')
plt.show()

# 将相似度矩阵保存到Excel文件
df = pd.DataFrame(similarity_matrix, index=labels, columns=labels)
excel_filename = f'{pwd}/similarity_matrix.xlsx'
df.to_excel(excel_filename, index=True)

# 对Excel文件应用条件格式，突出显示大于0.6的单元格
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# 加载工作簿和工作表
wb = load_workbook(excel_filename)
ws = wb.active

# 定义填充样式（黄色填充）
fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 定义数据范围
start_row = 2  # 数据从第二行开始（第一行是列标签）
start_col = 2  # 数据从第二列开始（第一列是行索引）
end_row = num_data_points + 1  # 总行数（包含标题行）
end_col = num_data_points + 1  # 总列数（包含索引列）

# 获取最后一列的列字母
last_column = get_column_letter(end_col)

# 定义数据范围，例如：B2:AE31
data_range = f'{get_column_letter(start_col)}{start_row}:{last_column}{end_row}'

# 应用条件格式
ws.conditional_formatting.add(data_range,
    CellIsRule(operator='greaterThan', formula=['0.6'], fill=fill))

# 保存Excel文件
wb.save(excel_filename)